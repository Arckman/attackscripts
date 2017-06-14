#---coding:u8-----
import logging
import os
import re
import openpyxl

l=logging.DEBUG
#l=logging.INFO
logging.basicConfig(level=l,format='%(asctime)s %(levelname)s [%(filename)s:%(lineno)d] %(message)s',datefmt='%a, %d %b %Y %H:%M:%S')

#path=r"D:\security\题目\sha ctf\2017\新建文件夹\http response".decode('u8').encode('gbk')
path=r'D:\security\题目\sha ctf\2017\新建文件夹\requests.xlsx'.decode('u8').encode('gbk')
#path=r'./requests'
t1='41'*145
c1='e4146d4252bafb3b38212df186497a7479d5e95af4796e7573a65e6849952032e4146d4252bafb3b38212df186497a7479d5e95af4796e7573a65e6849952032e4146d4252bafb3b38212df186497a7479d5e95af4796e7573a65e6849952032e4146d4252bafb3b38212df186497a7479d5e95af4796e7573a65e6849952032e4146d4252bafb3b38212df186497a74799edb6fda5b44'
c2='af7d6f4240be9a2d31252290ef5b7e797dd7fc3be66d6d6766b5375a79b84d428964052355a9f53759403fe18b416f7067d9e948e17d7d147eae52605cf4505f947c0c3e33dc8a5d593424f58928484157f7c33bf0747c7112976d406bb14136eb1105'

def analyse(path):
    tmp={}
    for (p,d,files) in os.walk(path):
        for f in files:
            t=open(os.path.join(p,f)).read()
            if re.search('html',t,re.IGNORECASE):
                t='OK'
            m=re.search('[0-9a-f]+$',f,re.I)
            if m!=None:
                tmp[rc4decode(t1,c1,m.group())]=t
    logging.debug(tmp)
    writefile('result.xlsx',tmp)

def readfile(path):#------------------right
    global t1,c1
    wb=openpyxl.load_workbook(path)
    ws=wb.active
    h=ws.max_row
    logging.debug('H is: '+str(h))
    rs=''
    lastpos='1'
    last=''
    p=re.compile("\(CASE WHEN \(SELECT SUBSTR\(flag,(\d+),1\)  FROM secret_flag LIMIT 0,1\) = '(.)' THEN stock ELSE price END\)")
    for i in range(2,h+1):
        c=ws.cell(row=i,column=7)
        c=c.value.strip()
        c2=re.search('[0-9a-f]+ HTTP/1\.1$',c).group()[:-9]
        t2=rc4decode(t1,c1,c2)
        # if t2=="(CASE WHEN (SELECT SUBSTR(flag,38,1)  FROM secret_flag LIMIT 0,1) = '}' THEN stock ELSE price END)":#for test
        #     print c2
        #print t2
        #analyse flag{}
        m=p.search(t2)
        if m!=None:
            pos=m.group(1)
            if pos!=lastpos:
                rs+=last
                lastpos=pos
            last=m.group(2)
    print rs


        #write file-------error,there is unusual char
        #ws2.append([rc4decode(t1,c1,c2)])
    #wb2.save('result.xlsx')

def writefile(name,m):
    wb=openpyxl.Workbook()
    ws=wb.active
    i=1
    for item in m.items():
        ws.append(item)
    wb.save(name)


def rc4decode(text1,cipher1,cipher2):#input must be hex string,return plaintext(char string)
    l=len(cipher1)
    if len(cipher2)<l:
        l=len(cipher2)
    #logging.debug('Length is %d'%(l))
    r=''
    for i in range(0,l/2):
        t=ord(text1[i*2:i*2+2].decode('hex'))^ord(cipher1[i*2:i*2+2].decode('hex'))^ord(cipher2[i*2:i*2+2].decode('hex'))
        #logging.debug('Word is %s(%s)'%(chr(t),hex(t)))
        r+=chr(t)
        #logging.debug(r)
    #logging.info('Text is %s'%(r))
    return r

readfile(path)
#rc4decode(t1,c1,c2)
#analyse(path)
